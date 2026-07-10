"""Excel export — five sheets: Financials, Valuation Results, Verdict Breakdown,
DCF Sensitivity, Assumptions & Quality."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Shared helpers ────────────────────────────────────────────────────────────

def _hs():
    return {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="1F4E79"),
        "alignment": Alignment(horizontal="center"),
    }


def _apply(cell, **styles):
    for attr, val in styles.items():
        setattr(cell, attr, val)


def _tb():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _fmt(val, decimals=2):
    if val is None:
        return "N/A"
    if isinstance(val, float):
        return round(val, decimals)
    return val


def _sec(ws, row, title, bg="D9E1F2"):
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=11)
    c.fill = PatternFill("solid", fgColor=bg)
    return c


def _row3(ws, r, label, value, unit):
    ws.cell(row=r, column=1, value=label).border = _tb()
    ws.cell(row=r, column=2, value=value).border = _tb()
    ws.cell(row=r, column=3, value=unit).border = _tb()


def _compute_risk_flags(f):
    """Mirror the JS evalShareholdingFlags logic using available server-side data."""
    flags = []
    pledged = f.get("promoter_pledged")
    if pledged is not None and pledged > 30:
        flags.append(
            f"High promoter pledging ({pledged:.1f}%) — forced selling risk in a downturn"
        )
    holding = f.get("promoter_holding")
    if holding is not None and holding < 30:
        flags.append(
            f"Low promoter skin-in-the-game ({holding:.1f}%) — management incentives"
            " may not align with minority shareholders"
        )
    return flags


# ── Main entry point ──────────────────────────────────────────────────────────

def build_excel(
    financials: dict,
    assumptions: dict,
    dcf: dict,
    rev_growth: dict,
    ps: dict,
    final_label: str,
    ticker: str,
    scorecard: dict = None,
    scenarios_matrix: list = None,
    g1_rates: list = None,
    tg_rates: list = None,
    pe_result: dict = None,
    governance: dict = None,
) -> bytes:
    pe_result = pe_result or {}
    wb = Workbook()
    hs = _hs()

    # ── Sheet 1: Financials ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Financials"
    for col, h in enumerate(["Field", "Value", "Unit"], 1):
        c = ws1.cell(row=1, column=col, value=h)
        _apply(c, **hs)
        c.border = _tb()

    fin_rows = [
        ("Company",                  financials.get("company_name"),                ""),
        ("Ticker",                   ticker.upper(),                                ""),
        ("Market Cap",               _fmt(financials.get("market_cap")),            "Cr"),
        ("Total Debt",               _fmt(financials.get("debt")),                  "Cr"),
        ("Cash & Equivalents",       _fmt(financials.get("cash")),                  "Cr"),
        ("Minority Interest",        _fmt(financials.get("minority_interest")),     "Cr"),
        ("Enterprise Value (EV)",    _fmt(financials.get("enterprise_value")),      "Cr"),
        ("Current Share Price",      _fmt(financials.get("share_price")),           "INR"),
        ("Revenue (Latest FY)",      _fmt(financials.get("revenue")),               "Cr"),
        ("EBITDA (Latest FY)",       _fmt(financials.get("ebitda")),                "Cr"),
        ("Net Profit (Latest FY)",   _fmt(financials.get("net_profit")),            "Cr"),
        ("EPS (Basic)",              _fmt(financials.get("eps")),                   "INR"),
        ("Shares Outstanding",       _fmt(financials.get("shares_outstanding"), 0), ""),
        ("5Y Avg EV/Revenue",        _fmt(financials.get("five_yr_avg_ev_rev")),    "x"),
        ("Years of Data Available",  financials.get("years_listed", "N/A"),         "yrs"),
        ("ROCE (Latest Year)",       _fmt(financials.get("roce_latest")),           "%"),
        ("ROCE (5Y Average)",        _fmt(financials.get("roce_5yr_avg")),          "%"),
        ("ROE (Latest Year)",        _fmt(financials.get("roe_latest")),            "%"),
        ("ROE (5Y Average)",         _fmt(financials.get("roe_5yr_avg")),           "%"),
        ("Net Profit Margin",        _fmt(financials.get("net_margin")),            "%"),
        ("Promoter Holding",         _fmt(financials.get("promoter_holding")),      "%"),
        ("Promoter Pledged",         _fmt(financials.get("promoter_pledged")),      "%"),
        ("FII Holding",              _fmt(financials.get("fii_holding")),           "%"),
        ("DII Holding",              _fmt(financials.get("dii_holding")),           "%"),
    ]
    for r, (field, value, unit) in enumerate(fin_rows, 2):
        _row3(ws1, r, field, value, unit)

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 10

    # ── Sheet 2: Valuation Results ────────────────────────────────────────────
    ws2 = wb.create_sheet("Valuation Results")

    r = 1
    _sec(ws2, r, "Method 1 — DCF Valuation")
    r += 1
    for label, val, unit in [
        ("Intrinsic Value per Share", _fmt(dcf.get("intrinsic_per_share")), "INR"),
        ("DCF Upside / Downside",     _fmt(dcf.get("upside_pct")),          "%"),
        ("PV of FCFs",                _fmt(dcf.get("sum_pv_fcf")),           "Cr"),
        ("PV of Terminal Value",      _fmt(dcf.get("pv_terminal")),          "Cr"),
        ("Net Debt",                  _fmt(dcf.get("net_debt")),             "Cr"),
        ("Equity Value",              _fmt(dcf.get("equity_value")),         "Cr"),
        ("Verdict",                   dcf.get("verdict"),                    ""),
    ]:
        _row3(ws2, r, label, val, unit)
        r += 1

    r += 1
    _sec(ws2, r, "Method 2 — Reverse Growth Check")
    r += 1
    for label, val, unit in [
        ("Implied Year 10 Revenue", _fmt(rev_growth.get("implied_rev_yr10")), "Cr"),
        ("Implied CAGR",            _fmt(rev_growth.get("implied_cagr")),      "%"),
        ("Your Assumed CAGR",       _fmt(rev_growth.get("your_cagr")),         "%"),
        ("Gap (Implied - Yours)",   _fmt(rev_growth.get("gap")),               "pp"),
        ("Verdict",                 rev_growth.get("verdict"),                 ""),
    ]:
        _row3(ws2, r, label, val, unit)
        r += 1

    r += 1
    _sec(ws2, r, "Method 3 — Historical EV/Revenue")
    r += 1
    for label, val, unit in [
        ("Current EV/Revenue",    _fmt(ps.get("current_ev_rev")),    "x"),
        ("5Y Average EV/Revenue", _fmt(ps.get("five_yr_avg_ev_rev")), "x"),
        ("Current vs Avg Ratio",  _fmt(ps.get("ratio_to_avg")),       "x"),
        ("Verdict",               ps.get("verdict"),                  ""),
    ]:
        _row3(ws2, r, label, val, unit)
        r += 1

    r += 1
    _sec(ws2, r, "Method 4 — Sector P/E Fair Value")
    r += 1
    for label, val, unit in [
        ("Intrinsic Value per Share", _fmt(pe_result.get("intrinsic_per_share")), "INR"),
        ("Upside / Downside",          _fmt(pe_result.get("upside_pct")),          "%"),
        ("EV/EBITDA",                  _fmt(pe_result.get("ev_ebitda")),            "x"),
        ("Verdict",                    pe_result.get("verdict"),                    ""),
    ]:
        _row3(ws2, r, label, val, unit)
        r += 1

    r += 2
    c = ws2.cell(row=r, column=1, value=f"FINAL VERDICT: {final_label}")
    c.font = Font(bold=True, size=14)
    c.fill = PatternFill("solid", fgColor="FFD700")

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 10

    # ── Sheet 3: Verdict Breakdown ────────────────────────────────────────────
    ws3 = wb.create_sheet("Verdict Breakdown")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 10
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 14

    METHOD_NAMES = {
        "dcf":        "DCF on Earnings",
        "rev_growth": "Reverse Growth Check",
        "ev_rev":     "Historical EV/Revenue",
        "pe":         "Sector P/E Fair Value",
    }

    if scorecard:
        r3 = 1
        rel       = scorecard.get("reliability_label", "N/A")
        rel_score = scorecard.get("reliability_score")
        rel_str   = f"{rel} (score: {_fmt(rel_score, 2)})" if rel_score is not None else rel
        ws3.cell(row=r3, column=1, value="DCF Reliability").font = Font(bold=True)
        ws3.cell(row=r3, column=2, value=rel_str)
        r3 += 2

        for col, h in enumerate(["Method", "Signal", "Weight", "Score", "Contribution"], 1):
            c = ws3.cell(row=r3, column=col, value=h)
            _apply(c, **hs)
            c.border = _tb()
        r3 += 1

        for method, mc in scorecard.get("contributions", {}).items():
            sig      = mc.get("signal_score")
            sig_str  = "—" if sig is None else ("+1" if sig > 0 else ("−1" if sig < 0 else "0"))
            w        = mc.get("weight", 0)
            contrib  = mc.get("contribution")
            c_str    = "—" if contrib is None else f"{'+' if contrib >= 0 else ''}{contrib:.4f}"

            ws3.cell(row=r3, column=1, value=METHOD_NAMES.get(method, method)).border = _tb()
            ws3.cell(row=r3, column=2, value=mc.get("verdict", "")).border = _tb()
            ws3.cell(row=r3, column=3, value=f"{w * 100:.0f}%").border = _tb()
            ws3.cell(row=r3, column=4, value=sig_str).border = _tb()
            ws3.cell(row=r3, column=5, value=c_str).border = _tb()
            r3 += 1

        r3 += 1
        ws = scorecard.get("weighted_score")
        ws3.cell(row=r3, column=1, value="Final Weighted Score").font = Font(bold=True)
        ws3.cell(row=r3, column=2, value=(
            f"{'+' if ws and ws >= 0 else ''}{ws:.3f}" if ws is not None else "N/A"
        ))
        r3 += 1
        ws3.cell(row=r3, column=1, value="Final Verdict").font = Font(bold=True)
        final_cell = ws3.cell(row=r3, column=2, value=scorecard.get("final_label", final_label))
        final_cell.font = Font(bold=True, size=12)
    else:
        ws3.cell(row=1, column=1, value="Scorecard data not available — run a valuation first")

    # ── Sheet 4: DCF Sensitivity ──────────────────────────────────────────────
    ws4 = wb.create_sheet("DCF Sensitivity")

    GREEN_FILL  = PatternFill("solid", fgColor="DCFCE7")
    RED_FILL    = PatternFill("solid", fgColor="FEE2E2")
    YELLOW_FILL = PatternFill("solid", fgColor="FEF9C3")

    if scenarios_matrix and g1_rates and tg_rates:
        # Header row
        corner = ws4.cell(row=1, column=1, value="TGR ↓ / G1 →")
        corner.font = Font(bold=True)
        corner.fill = PatternFill("solid", fgColor="D9E1F2")
        corner.border = _tb()

        for ci, g1 in enumerate(g1_rates, 2):
            c = ws4.cell(row=1, column=ci, value=f"{g1}%/yr")
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F4E79")
            c.alignment = Alignment(horizontal="center")
            c.border = _tb()

        for ri, (tg, row_data) in enumerate(zip(tg_rates, scenarios_matrix), 2):
            lbl = ws4.cell(row=ri, column=1, value=f"{tg}%")
            lbl.font = Font(bold=True)
            lbl.fill = PatternFill("solid", fgColor="D9E1F2")
            lbl.border = _tb()
            lbl.alignment = Alignment(horizontal="center")

            for ci, cell_data in enumerate(row_data, 2):
                val = cell_data.get("intrinsic_per_share")
                up  = cell_data.get("upside_pct")
                txt = f"₹{round(val):,}" if val is not None else "N/A"
                xc  = ws4.cell(row=ri, column=ci, value=txt)
                xc.alignment = Alignment(horizontal="center")
                xc.border = _tb()
                if val is not None:
                    if up is not None and up > 20:
                        xc.fill = GREEN_FILL
                    elif up is not None and up < -20:
                        xc.fill = RED_FILL
                    else:
                        xc.fill = YELLOW_FILL

        ws4.column_dimensions["A"].width = 14
        for ci in range(2, len(g1_rates) + 2):
            ws4.column_dimensions[get_column_letter(ci)].width = 14
    else:
        ws4.cell(row=1, column=1, value="Sensitivity matrix data not available — run a valuation first")

    # ── Sheet 5: Assumptions & Quality ───────────────────────────────────────
    ws5 = wb.create_sheet("Assumptions & Quality")
    ws5.column_dimensions["A"].width = 32
    ws5.column_dimensions["B"].width = 18
    ws5.column_dimensions["C"].width = 10

    r5 = 1

    # — Business Quality section —
    _sec(ws5, r5, "Business Quality Metrics")
    r5 += 1
    for col, h in enumerate(["Metric", "Value", "Unit"], 1):
        c = ws5.cell(row=r5, column=col, value=h)
        _apply(c, **hs)
        c.border = _tb()
    r5 += 1
    for label, val, unit in [
        ("ROCE (Latest Year)",  _fmt(financials.get("roce_latest")),  "%"),
        ("ROCE (5Y Average)",   _fmt(financials.get("roce_5yr_avg")), "%"),
        ("ROE (Latest Year)",   _fmt(financials.get("roe_latest")),   "%"),
        ("ROE (5Y Average)",    _fmt(financials.get("roe_5yr_avg")),  "%"),
        ("Net Profit Margin",   _fmt(financials.get("net_margin")),   "%"),
    ]:
        _row3(ws5, r5, label, val, unit)
        r5 += 1

    # — Shareholding section —
    r5 += 1
    _sec(ws5, r5, "Shareholding")
    r5 += 1
    for col, h in enumerate(["Metric", "Value", "Unit"], 1):
        c = ws5.cell(row=r5, column=col, value=h)
        _apply(c, **hs)
        c.border = _tb()
    r5 += 1
    for label, val, unit in [
        ("Promoter Holding",  _fmt(financials.get("promoter_holding")),  "%"),
        ("Promoter Pledged",  _fmt(financials.get("promoter_pledged")),  "%"),
        ("FII Holding",       _fmt(financials.get("fii_holding")),       "%"),
        ("DII Holding",       _fmt(financials.get("dii_holding")),       "%"),
    ]:
        _row3(ws5, r5, label, val, unit)
        r5 += 1

    # — Active Risk Flags section —
    r5 += 1
    _sec(ws5, r5, "Active Risk Flags")
    r5 += 1
    flags = _compute_risk_flags(financials)
    if flags:
        for flag in flags:
            fc = ws5.cell(row=r5, column=1, value=f"⚠ {flag}")
            fc.font = Font(color="B91C1C")
            r5 += 1
    else:
        ok = ws5.cell(row=r5, column=1, value="No active risk flags detected")
        ok.font = Font(color="15803D")
        r5 += 1

    # — Assumptions section —
    r5 += 1
    _sec(ws5, r5, "Valuation Assumptions")
    r5 += 1
    for col, h in enumerate(["Parameter", "Value", "Unit"], 1):
        c = ws5.cell(row=r5, column=col, value=h)
        _apply(c, **hs)
        c.border = _tb()
    r5 += 1
    for label, val, unit in [
        ("Revenue Growth Yr 1-5",       assumptions.get("growth_rate_1_5"),  "%"),
        ("Revenue Growth Yr 6-10",      assumptions.get("growth_rate_6_10"), "%"),
        ("Target Net Margin Yr 10",     assumptions.get("target_margin"),     "%"),
        ("Discount Rate (WACC)",        assumptions.get("wacc"),              "%"),
        ("Terminal Growth Rate",        assumptions.get("terminal_rate"),     "%"),
        ("Sector EV/Revenue Multiple",  assumptions.get("sector_ev_rev"),     "x"),
        ("Sector P/E Multiple",         assumptions.get("sector_pe"),         "x"),
    ]:
        _row3(ws5, r5, label, val, unit)
        r5 += 1

    # ── Sheet 6: Governance (only when governance data is present) ───────────
    if governance:
        ws6 = wb.create_sheet("Governance")
        ws6.column_dimensions["A"].width = 36
        ws6.column_dimensions["B"].width = 28
        r6 = 1

        # Title row
        th = ws6.cell(row=r6, column=1, value=f"Governance — {ticker.upper()}")
        th.font = Font(bold=True, size=13, color="1F4E79")
        r6 += 1

        # Metadata
        _sec(ws6, r6, "Fetch Metadata", bg="EFF6FF")
        r6 += 1
        src = governance.get("governance_source", "lightweight")
        src_label = "Auto-fetched (lightweight)" if src == "lightweight" else "Deep Research"
        fetched_at = governance.get("data_fetched_at") or "—"
        fp = governance.get("fields_populated", 0)
        ft = governance.get("fields_total", 34)
        dur = governance.get("fetch_duration_seconds")
        for lbl, val in [
            ("Source",           src_label),
            ("Fields populated", f"{fp} / {ft}"),
            ("Fetch duration",   f"{dur}s" if dur is not None else "—"),
            ("Fetched at",       str(fetched_at)[:19] if fetched_at != "—" else "—"),
        ]:
            ws6.cell(row=r6, column=1, value=lbl).border = _tb()
            ws6.cell(row=r6, column=2, value=val).border = _tb()
            r6 += 1

        # Risk flags
        r6 += 1
        _sec(ws6, r6, "Governance Risk Flags", bg="FEE2E2")
        r6 += 1
        for col, h in enumerate(["Flag", "Status"], 1):
            c = ws6.cell(row=r6, column=col, value=h)
            _apply(c, **hs)
            c.border = _tb()
        r6 += 1
        flag_defs = [
            ("flag_auditor_changed",            "Auditor Change"),
            ("flag_caro_qualified",             "CARO Qualification"),
            ("flag_promoter_pledging_high",     "High Promoter Pledging (>30%)"),
            ("flag_promoter_stake_declining",   "Promoter Stake Declining"),
            ("flag_working_capital_deteriorating", "Working Capital Deteriorating"),
        ]
        for key, label in flag_defs:
            val = governance.get(key)
            if val is True:
                status, color = "⚠ Yes", "B91C1C"
            elif val is False:
                status, color = "✓ No", "15803D"
            else:
                status, color = "— Not assessed", "64748B"
            ws6.cell(row=r6, column=1, value=label).border = _tb()
            vc = ws6.cell(row=r6, column=2, value=status)
            vc.border = _tb()
            vc.font = Font(color=color)
            r6 += 1

        # Shareholding
        r6 += 1
        _sec(ws6, r6, "Shareholding (latest quarter)", bg="DCFCE7")
        r6 += 1
        for lbl, key in [
            ("Promoter Holding %",  "promoter_holding_pct"),
            ("FII Holding %",       "fii_holding_pct"),
            ("DII Holding %",       "dii_holding_pct"),
            ("Promoter Pledged %",  "promoter_pledged_pct"),
            ("Promoter QoQ change", "promoter_holding_change_qoq"),
            ("FII QoQ change",      "fii_change_qoq"),
        ]:
            v = governance.get(key)
            ws6.cell(row=r6, column=1, value=lbl).border = _tb()
            ws6.cell(row=r6, column=2, value=f"{v:.2f}%" if isinstance(v, (int, float)) else "—").border = _tb()
            r6 += 1

        # Promoter holding trend
        trend = governance.get("promoter_holding_trend") or []
        if any(v is not None for v in trend):
            ws6.cell(row=r6, column=1, value="Promoter Trend (8Q, old→new)").border = _tb()
            ws6.cell(row=r6, column=2, value=" | ".join(
                f"{v:.1f}%" if v is not None else "—" for v in trend
            )).border = _tb()
            r6 += 1

        # Dividend history
        r6 += 1
        _sec(ws6, r6, "Dividend History", bg="FEF9C3")
        r6 += 1
        dps = governance.get("dividend_per_share_5yr") or []
        for lbl, key in [
            ("Dividend Yield %",    "dividend_yield_latest"),
            ("Consistency",         "dividend_consistency"),
            ("Growing",             "dividend_growing"),
        ]:
            v = governance.get(key)
            if isinstance(v, bool):
                disp = "Yes" if v else "No"
            elif isinstance(v, (int, float)):
                disp = f"{v:.2f}%"
            else:
                disp = "—"
            ws6.cell(row=r6, column=1, value=lbl).border = _tb()
            ws6.cell(row=r6, column=2, value=disp).border = _tb()
            r6 += 1
        if dps:
            ws6.cell(row=r6, column=1, value="DPS (₹) 5Y (old→new)").border = _tb()
            ws6.cell(row=r6, column=2, value=" | ".join(
                f"{v:.2f}" if v is not None else "—" for v in dps
            )).border = _tb()
            r6 += 1

        # Working capital
        r6 += 1
        _sec(ws6, r6, "Working Capital Health", bg="EDE9FE")
        r6 += 1
        for lbl, key in [
            ("Latest CCC (days)",       "cash_conversion_cycle_latest"),
            ("WC Deteriorating",        "flag_working_capital_deteriorating"),
            ("Deterioration Reason",    "deterioration_reason"),
        ]:
            v = governance.get(key)
            if isinstance(v, bool):
                disp = "Yes" if v else "No"
            elif v is None:
                disp = "—"
            else:
                disp = str(v)
            ws6.cell(row=r6, column=1, value=lbl).border = _tb()
            ws6.cell(row=r6, column=2, value=disp).border = _tb()
            r6 += 1

        # Concall snippet
        snippet = governance.get("mgmt_commentary_snippet")
        if snippet:
            r6 += 1
            _sec(ws6, r6, "Management Commentary", bg="E0F2FE")
            r6 += 1
            src_lbl = governance.get("concall_source") or "Screener"
            ws6.cell(row=r6, column=1, value="Source").border = _tb()
            ws6.cell(row=r6, column=2, value=src_lbl).border = _tb()
            r6 += 1
            tc = ws6.cell(row=r6, column=1, value=snippet[:500])
            tc.alignment = Alignment(wrap_text=True)
            ws6.row_dimensions[r6].height = 80
            ws6.merge_cells(start_row=r6, start_column=1, end_row=r6, end_column=2)
            r6 += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
